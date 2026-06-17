from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.core.mail import EmailMessage
from django.core.paginator import Paginator
from django.contrib.auth import login
from django.contrib.auth.forms import UserCreationForm
from openpyxl import Workbook
from io import BytesIO

from .models import (
    Product, Category, Manufacturer, 
    Cart, CartItem, Order, OrderItem, Profile
)

from django.contrib.auth.views import PasswordChangeView, PasswordChangeDoneView
from django.urls import reverse_lazy

def index(request):
    popular_products = Product.objects.all().order_by('-id')[:6]
    categories = Category.objects.all()
    return render(request, 'shop/index.html', {
        'popular_products': popular_products,
        'categories': categories,
    })


def about(request):
    return render(request, 'shop/about.html')


def shop_info(request):
    return render(request, 'shop/shop_info.html')


def product_list(request):
    products = Product.objects.all().order_by('-id')
    category_id = request.GET.get('category')
    if category_id:
        products = products.filter(category_id=category_id)
    manufacturer_id = request.GET.get('manufacturer')
    if manufacturer_id:
        products = products.filter(manufacturer_id=manufacturer_id)
    
    search_query = request.GET.get('search')
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) | Q(description__icontains=search_query)
        )
    
    paginator = Paginator(products, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'shop/catalog.html', {
        'page_obj': page_obj,
        'categories': Category.objects.all(),
        'manufacturers': Manufacturer.objects.all(),
        'current_category': category_id,
        'current_manufacturer': manufacturer_id,
        'search_query': search_query,
    })


def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'shop/product_detail.html', {'product': product})


@login_required
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_item, created = CartItem.objects.get_or_create(
        cart=cart,
        product=product,
        defaults={'quantity': 1}
    )
    if not created:
        cart_item.quantity += 1
        cart_item.save()
    return redirect('shop:product_list')


@login_required
def update_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    if request.method == 'POST':
        new_quantity = int(request.POST.get('quantity'))
        if new_quantity <= cart_item.product.stock:
            if new_quantity > 0:
                cart_item.quantity = new_quantity
                cart_item.save()
            else:
                cart_item.delete()
    return redirect('shop:cart_view')


@login_required
def remove_from_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    cart_item.delete()
    return redirect('shop:cart_view')


@login_required
def cart_view(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_items = cart.cartitem_set.all()
    total_price = 0
    for item in cart_items:
        total_price += item.product.price * item.quantity
    return render(request, 'shop/cart.html', {
        'cart_items': cart_items,
        'total_price': total_price
    })


@login_required
def checkout(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_items = cart.cartitem_set.all()
    if not cart_items:
        return redirect('shop:product_list')
    
    if request.method != 'POST':
        total = sum(item.product.price * item.quantity for item in cart_items)
        return render(request, 'shop/checkout.html', {
            'cart_items': cart_items,
            'total': total
        })
    
    total = sum(item.product.price * item.quantity for item in cart_items)
    order = Order.objects.create(
        user=request.user,
        address=request.POST.get('address', ''),
        phone=request.POST.get('phone', ''),
        total=total
    )
    for item in cart_items:
        OrderItem.objects.create(
            order=order,
            product=item.product.name,
            quantity=item.quantity,
            price=item.product.price
        )
    
    wb = Workbook()
    ws = wb.active
    ws['A1'] = f'Чек заказа №{order.id}'
    ws['A2'] = f'Покупатель: {request.user.username}'
    ws['A3'] = f'Телефон: {request.POST.get("phone", "")}'
    ws['A4'] = f'Адрес: {request.POST.get("address", "")}'
    ws['A6'] = 'Товар'
    ws['B6'] = 'Кол-во'
    ws['C6'] = 'Цена'
    ws['D6'] = 'Сумма'
    
    row = 7
    for item in cart_items:
        ws[f'A{row}'] = item.product.name
        ws[f'B{row}'] = item.quantity
        ws[f'C{row}'] = float(item.product.price)
        ws[f'D{row}'] = float(item.product.price * item.quantity)
        row += 1
    
    ws[f'C{row+1}'] = 'ИТОГО:'
    ws[f'D{row+1}'] = float(total)
    
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    email = EmailMessage(
        subject=f'Чек заказа №{order.id}',
        body=f'Спасибо за покупку!\n\nВаш заказ №{order.id} на сумму {total} руб. оформлен.\nЧек во вложении.',
        from_email='shop@example.com',
        to=[request.user.email],
    )
    email.attach(
        'receipt.xlsx', 
        excel_file.getvalue(), 
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    email.send()
    
    cart_items.delete()
    return render(request, 'shop/order_success.html', {'order': order})

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            Profile.objects.get_or_create(user=user)
            login(request, user)
            return redirect('shop:index')
    else:
        form = UserCreationForm()
    return render(request, 'shop/register.html', {'form': form})


@login_required
def profile(request):
    orders = Order.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'shop/profile.html', {
        'user': request.user,
        'profile': request.user.profile,
        'orders': orders,
    })

@login_required
def settings_view(request):
    profile = request.user.profile
    if request.method == 'POST':
        profile.favorite_category_id = request.POST.get('favorite_category') or None
        profile.delivery_city = request.POST.get('delivery_city', '')
        profile.postal_code = request.POST.get('postal_code', '')
        profile.save()
        from django.contrib import messages
        messages.success(request, '✅ Настройки сохранены!')
        return redirect('shop:settings')
    
    categories = Category.objects.all()
    return render(request, 'shop/settings.html', {
        'profile': profile,
        'categories': categories,
    })

class CustomPasswordChangeView(PasswordChangeView):
    template_name = 'shop/settings_password.html'
    success_url = reverse_lazy('shop:password_change_done')


class CustomPasswordChangeDoneView(PasswordChangeDoneView):
    template_name = 'shop/settings_password_done.html'